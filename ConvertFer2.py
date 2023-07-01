import pandas as pd
import openpyxl
from openpyxl import Workbook , load_workbook
import csv
import Ferro2Excel
from openpyxl.styles.numbers import BUILTIN_FORMATS
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from datetime import datetime
from openpyxl.styles import Border, Side
from reportlab.lib.pagesizes import landscape
from pdf2image import convert_from_path
import os
from dotenv import load_dotenv
load_dotenv()
wb = Workbook()
ws = wb.active
ws.merge_cells('B1:D1')
ws['B1'] = 'Ферросплавы и руды'
ws['E1'] = 'Хром'
ws['A1'] = Ferro2Excel.DATA
ws['A1'].number_format
ws.merge_cells('A2:A3')
ws.merge_cells('B2:B3')
ws.merge_cells('C2:E2')
ws['A2'] = 'Продукция'
ws['B2'] = 'Цена'
ws['C2'] = 'Изменения относительно предыдущего периода'
ws['C3'] = 'День'
ws['D3'] = 'Неделя'
ws['E3'] = 'Месяц'
ws['A4'] = 'HC FeCr'
ws['A5'] = Ferro2Excel.cellFE
ws['A6'] = Ferro2Excel.cellUg
ws['A7'] = Ferro2Excel.cellKo
ws['A8'] = Ferro2Excel.cellLo
ws['A9'] = Ferro2Excel.cellRu
ws['A10'] = 'LC FeCr'
ws['A11'] = Ferro2Excel.cellIn
ws['A12'] = Ferro2Excel.cellCh
ws['A13'] = Ferro2Excel.cellFob
ws['A14'] = 'Cr руда'
ws['A15'] = Ferro2Excel.cellFR
ws['B5'] = str(Ferro2Excel.PRICE1) + " CNY/т"
ws['B6'] = str(Ferro2Excel.PRICE2) + " USDc/фунт Cr"
ws['B7'] = str(Ferro2Excel.PRICE3) + " USDc/фунт Cr"
ws['B8'] = str(Ferro2Excel.PRICE4) + " USDc/фунт Cr"
ws['B9'] = str(Ferro2Excel.PRICE5) + " USDc/фунт Cr"
ws['B11'] = str(Ferro2Excel.PRICE6) + " CNY/т"
ws['B12'] = str(Ferro2Excel.PRICE7) + " USDc/фунт Cr"
ws['B13'] = str(Ferro2Excel.PRICE8) + " USDc/фунт Cr"
ws['B15'] = str(Ferro2Excel.PRICE9) + " USD/т"
ws['C5'] = f"{Ferro2Excel.CH_DAY1:+}" + '  (' + f"{Ferro2Excel.CH_DAY_PR1:+}" + " %)"
ws['C6'] = f"{Ferro2Excel.CH_DAY2:+}" + '  (' + f"{Ferro2Excel.CH_DAY_PR2:+}" + " %)"
ws['C7'] = f"{Ferro2Excel.CH_DAY3:+}" + '  (' + f"{Ferro2Excel.CH_DAY_PR3:+}" + " %)"
ws['C8'] = f"{Ferro2Excel.CH_DAY4:+}" + '  (' + f"{Ferro2Excel.CH_DAY_PR4:+}" + " %)"
ws['C9'] = f"{Ferro2Excel.CH_DAY5:+}" + '  (' + f"{Ferro2Excel.CH_DAY_PR5:+}" + " %)"
ws['C11'] = f"{Ferro2Excel.CH_DAY6:+}" + '  (' + f"{Ferro2Excel.CH_DAY_PR6:+}" + " %)"
ws['C12'] = f"{Ferro2Excel.CH_DAY7:+}" + '  (' + f"{Ferro2Excel.CH_DAY_PR7:+}" + " %)"
ws['C13'] = f"{Ferro2Excel.CH_DAY8:+}" + '  (' + f"{Ferro2Excel.CH_DAY_PR8:+}" + " %)"
ws['C15'] = f"{Ferro2Excel.CH_DAY9:+}" + '  (' + f"{Ferro2Excel.CH_DAY_PR9:+}" + " %)"
ws['D5'] = f"{Ferro2Excel.CH_W1:+}" + '  (' + f"{Ferro2Excel.CH_W_PR1:+}" + " %)"
ws['D6'] = f"{Ferro2Excel.CH_W2:+}" + '  (' + f"{Ferro2Excel.CH_W_PR2:+}" + " %)"
ws['D7'] = f"{Ferro2Excel.CH_W3:+}" + '  (' + f"{Ferro2Excel.CH_W_PR3:+}" + " %)"
ws['D8'] = f"{Ferro2Excel.CH_W4:+}" + '  (' + f"{Ferro2Excel.CH_W_PR4:+}" + " %)"
ws['D9'] = f"{Ferro2Excel.CH_W5:+}" + '  (' + f"{Ferro2Excel.CH_W_PR5:+}" + " %)"
ws['D11'] = f"{Ferro2Excel.CH_W6:+}" + '  (' + f"{Ferro2Excel.CH_W_PR6:+}" + " %)"
ws['D12'] = f"{Ferro2Excel.CH_W7:+}" + '  (' + f"{Ferro2Excel.CH_W_PR7:+}" + " %)"
ws['D13'] = f"{Ferro2Excel.CH_W8:+}" + '  (' + f"{Ferro2Excel.CH_W_PR8:+}" + " %)"
ws['D15'] = f"{Ferro2Excel.CH_W9:+}" + '  (' + f"{Ferro2Excel.CH_W_PR9:+}" + " %)"
ws['E5'] = f"{Ferro2Excel.CH_M1:+}" + '  (' + f"{Ferro2Excel.CH_M_PR1:+}" + " %)"
ws['E6'] = f"{Ferro2Excel.CH_M2:+}" + '  (' + f"{Ferro2Excel.CH_M_PR2:+}" + " %)"
ws['E7'] = f"{Ferro2Excel.CH_M3:+}" + '  (' + f"{Ferro2Excel.CH_M_PR3:+}" + " %)"
ws['E8'] = f"{Ferro2Excel.CH_M4:+}" + '  (' + f"{Ferro2Excel.CH_M_PR4:+}" + " %)"
ws['E9'] = f"{Ferro2Excel.CH_M5:+}" + '  (' + f"{Ferro2Excel.CH_M_PR5:+}" + " %)"
ws['E11'] = f"{Ferro2Excel.CH_M6:+}" + '  (' + f"{Ferro2Excel.CH_M_PR6:+}" + " %)"
ws['E12'] = f"{Ferro2Excel.CH_M7:+}" + '  (' + f"{Ferro2Excel.CH_M_PR7:+}" + " %)"
ws['E13'] = f"{Ferro2Excel.CH_M8:+}" + '  (' + f"{Ferro2Excel.CH_M_PR8:+}" + " %)"
ws['E15'] = f"{Ferro2Excel.CH_M9:+}" + '  (' + f"{Ferro2Excel.CH_M_PR9:+}" + " %)"
def set_border(ws, cell_range):
    thin = Side(border_style="thin", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
set_border(ws, 'A1:E15') 
wb.save('Fer2.xlsx')
wb = openpyxl.load_workbook('Fer2.xlsx')
sheet = wb.active
custom_page_width = 1520 
custom_page_height = 340
custom_page_size = (custom_page_width, custom_page_height)
c = canvas.Canvas('Fer2.pdf', pagesize=custom_page_size)
pdfmetrics.registerFont(TTFont('Arial', 'arial.ttf'))
num_rows = sheet.max_row
num_cols = sheet.max_column
cell_width = 300
cell_height = 20
x_start = 10
y_start = 320
for i in range(1, num_rows + 1):
    y = y_start - (i * cell_height)
    for j in range(1, num_cols + 1):
        x = x_start + ((j - 1) * cell_width)
        value = sheet.cell(row=i, column=j).value
        if isinstance(value, (float, int)):
            value = str(value)
        elif isinstance(value, datetime):
            value = value.strftime('%Y-%m-%d %H:%M:%S')
        elif value is None:
            value = ''
        c.setFont('Arial', 10)
        c.drawCentredString(x + (cell_width / 2), y + (cell_height / 2), value)
        c.rect(x, y, cell_width, cell_height)
    for i in range(num_rows + 1):
        y = y_start - (i * cell_height)
        c.line(x_start, y, x_start + (num_cols * cell_width), y)
    for j in range(num_cols + 1):
        x = x_start + (j * cell_width)
        c.line(x, y_start, x, y_start - (num_rows * cell_height))
c.save()
images = convert_from_path('Fer2.pdf', 700,poppler_path=os.getenv('POPPLERSUPER'))
for image in images:
    image.save('Fer2.png')