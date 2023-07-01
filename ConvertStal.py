import pandas as pd
import openpyxl
from openpyxl import Workbook , load_workbook
import csv
import StalnaiaExcel
from openpyxl.styles.numbers import BUILTIN_FORMATS
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from datetime import datetime
from openpyxl.styles import Border, Side
from reportlab.lib.pagesizes import landscape
from pdf2image import convert_from_path
import os
from dotenv import load_dotenv
load_dotenv()
wb = Workbook()
ws = wb.active
ws.merge_cells('B1:E1')
ws['B1'] = 'Стальная продукция'
ws['A1'] = StalnaiaExcel.DATA
ws['A1'].number_format
ws.merge_cells('A2:A3')
ws.merge_cells('B2:B3')
ws.merge_cells('C2:E2')
ws['A2'] = 'Продукция'
ws['B2'] = 'Цена'
ws['C2'] = 'Изменения относительно предыдущего периода'
ws['C3'] = 'День'
ws['D3'] = 'Неделя'
ws['E3'] = 'Месяц'
ws['A4'] = 'Заготовка'
ws['A5'] = StalnaiaExcel.cellFE
ws['A6'] = StalnaiaExcel.cellUg
ws['A7'] = 'Рулон г/к'
ws['A8'] = StalnaiaExcel.cellKo
ws['A9'] = StalnaiaExcel.cellLo
ws['A10'] = StalnaiaExcel.cellRu
ws['A11'] = 'Рулон х/к'
ws['A12'] = StalnaiaExcel.cellIn
ws['A13'] = StalnaiaExcel.cellCh
ws['A14'] = StalnaiaExcel.cellFob
ws['A15'] = 'Арматура'
ws['A16'] = StalnaiaExcel.cellFR
ws['A17'] = StalnaiaExcel.cellFAC
ws['B5'] = str(StalnaiaExcel.PRICE1) + " USD/т"
ws['B6'] = str(StalnaiaExcel.PRICE2) + " USD/т"
ws['B8'] = str(StalnaiaExcel.PRICE3) + " USD/т"
ws['B9'] = str(StalnaiaExcel.PRICE4) + " USD/т"
ws['B10'] = str(StalnaiaExcel.PRICE5) + " руб/т"
ws['B12'] = str(StalnaiaExcel.PRICE6) + " USD/т"
ws['B13'] = str(StalnaiaExcel.PRICE7) + " USD/т"
ws['B14'] = str(StalnaiaExcel.PRICE8) + " руб/т"
ws['B16'] = str(StalnaiaExcel.PRICE9) + " USD/т"
ws['B17'] = str(StalnaiaExcel.PRICE10) + " руб/т"
ws['C5'] = f"{StalnaiaExcel.CH_DAY1:+}" + '  (' + f"{StalnaiaExcel.CH_DAY_PR1:+}" + " %)"
ws['C6'] = f"{StalnaiaExcel.CH_DAY2:+}" + '  (' + f"{StalnaiaExcel.CH_DAY_PR2:+}" + " %)"
ws['C8'] = f"{StalnaiaExcel.CH_DAY3:+}" + '  (' + f"{StalnaiaExcel.CH_DAY_PR3:+}" + " %)"
ws['C9'] = f"{StalnaiaExcel.CH_DAY4:+}" + '  (' + f"{StalnaiaExcel.CH_DAY_PR4:+}" + " %)"
ws['C10'] = f"{StalnaiaExcel.CH_DAY5:+}" + '  (' + f"{StalnaiaExcel.CH_DAY_PR5:+}" + " %)"
ws['C12'] = f"{StalnaiaExcel.CH_DAY6:+}" + '  (' + f"{StalnaiaExcel.CH_DAY_PR6:+}" + " %)"
ws['C13'] = f"{StalnaiaExcel.CH_DAY7:+}" + '  (' + f"{StalnaiaExcel.CH_DAY_PR7:+}" + " %)"
ws['C14'] = f"{StalnaiaExcel.CH_DAY8:+}" + '  (' + f"{StalnaiaExcel.CH_DAY_PR8:+}" + " %)"
ws['C16'] = f"{StalnaiaExcel.CH_DAY9:+}" + '  (' + f"{StalnaiaExcel.CH_DAY_PR9:+}" + " %)"
ws['C17'] = f"{StalnaiaExcel.CH_DAY10:+}" + '  (' + f"{StalnaiaExcel.CH_DAY_PR10:+}" + " %)"
ws['D5'] = f"{StalnaiaExcel.CH_W1:+}" + '  (' + f"{StalnaiaExcel.CH_W_PR1:+}" + " %)"
ws['D6'] = f"{StalnaiaExcel.CH_W2:+}" + '  (' + f"{StalnaiaExcel.CH_W_PR2:+}" + " %)"
ws['D8'] = f"{StalnaiaExcel.CH_W3:+}" + '  (' + f"{StalnaiaExcel.CH_W_PR3:+}" + " %)"
ws['D9'] = f"{StalnaiaExcel.CH_W4:+}" + '  (' + f"{StalnaiaExcel.CH_W_PR4:+}" + " %)"
ws['D10'] = f"{StalnaiaExcel.CH_W5:+}" + '  (' + f"{StalnaiaExcel.CH_W_PR5:+}" + " %)"
ws['D12'] = f"{StalnaiaExcel.CH_W6:+}" + '  (' + f"{StalnaiaExcel.CH_W_PR6:+}" + " %)"
ws['D13'] = f"{StalnaiaExcel.CH_W7:+}" + '  (' + f"{StalnaiaExcel.CH_W_PR7:+}" + " %)"
ws['D14'] = f"{StalnaiaExcel.CH_W8:+}" + '  (' + f"{StalnaiaExcel.CH_W_PR8:+}" + " %)"
ws['D16'] = f"{StalnaiaExcel.CH_W9:+}" + '  (' + f"{StalnaiaExcel.CH_W_PR9:+}" + " %)"
ws['D17'] = f"{StalnaiaExcel.CH_W10:+}" + '  (' + f"{StalnaiaExcel.CH_W_PR10:+}" + " %)"
ws['E5'] = f"{StalnaiaExcel.CH_M1:+}" + '  (' + f"{StalnaiaExcel.CH_M_PR1:+}" + " %)"
ws['E6'] = f"{StalnaiaExcel.CH_M2:+}" + '  (' + f"{StalnaiaExcel.CH_M_PR2:+}" + " %)"
ws['E8'] = f"{StalnaiaExcel.CH_M3:+}" + '  (' + f"{StalnaiaExcel.CH_M_PR3:+}" + " %)"
ws['E9'] = f"{StalnaiaExcel.CH_M4:+}" + '  (' + f"{StalnaiaExcel.CH_M_PR4:+}" + " %)"
ws['E10'] = f"{StalnaiaExcel.CH_M5:+}" + '  (' + f"{StalnaiaExcel.CH_M_PR5:+}" + " %)"
ws['E12'] = f"{StalnaiaExcel.CH_M6:+}" + '  (' + f"{StalnaiaExcel.CH_M_PR6:+}" + " %)"
ws['E13'] = f"{StalnaiaExcel.CH_M7:+}" + '  (' + f"{StalnaiaExcel.CH_M_PR7:+}" + " %)"
ws['E14'] = f"{StalnaiaExcel.CH_M8:+}" + '  (' + f"{StalnaiaExcel.CH_M_PR8:+}" + " %)"
ws['E16'] = f"{StalnaiaExcel.CH_M9:+}" + '  (' + f"{StalnaiaExcel.CH_M_PR9:+}" + " %)"
ws['E17'] = f"{StalnaiaExcel.CH_M10:+}" + '  (' + f"{StalnaiaExcel.CH_M_PR10:+}" + " %)"
def set_border(ws, cell_range):
    thin = Side(border_style="thin", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
set_border(ws, 'A1:E17') 
wb.save('Stal.xlsx')
wb = openpyxl.load_workbook('Stal.xlsx')
sheet = wb.active
custom_page_width = 1270 
custom_page_height = 380
custom_page_size = (custom_page_width, custom_page_height)
c = canvas.Canvas('Stal.pdf', pagesize=custom_page_size)
pdfmetrics.registerFont(TTFont('Arial', 'arial.ttf'))
num_rows = sheet.max_row
num_cols = sheet.max_column
cell_width = 250
cell_height = 20
x_start = 10
y_start = 360
for i in range(1, num_rows + 1):
    y = y_start - (i * cell_height)
    for j in range(1, num_cols + 1):
        x = x_start + ((j - 1) * cell_width)
        value = sheet.cell(row=i, column=j).value
        if isinstance(value, (float, int)):
            value = str(value)
        elif isinstance(value, datetime):
            value = value.strftime('%Y-%m-%d %H:%M:%S')
        elif value is None:
            value = ''
        c.setFont('Arial', 10)
        c.drawCentredString(x + (cell_width / 2), y + (cell_height / 2), value)
        c.rect(x, y, cell_width, cell_height)
    for i in range(num_rows + 1):
        y = y_start - (i * cell_height)
        c.line(x_start, y, x_start + (num_cols * cell_width), y)
    for j in range(num_cols + 1):
        x = x_start + (j * cell_width)
        c.line(x, y_start, x, y_start - (num_rows * cell_height))
c.save()
images = convert_from_path('Stal.pdf', 700,poppler_path=os.getenv('POPPLERSUPER'))
for image in images:
    image.save('Stal.png')