import pandas as pd
import openpyxl
from openpyxl import Workbook , load_workbook
import csv
import SurevyeExcel
from openpyxl.styles.numbers import BUILTIN_FORMATS
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from datetime import datetime
from openpyxl.styles import Border, Side
from reportlab.lib.pagesizes import landscape
from pdf2image import convert_from_path
import os
from dotenv import load_dotenv
load_dotenv()
wb = Workbook()
ws = wb.active
ws.merge_cells('B1:E1')
ws['B1'] = 'Сырьевые материалы'
ws['A1'] = SurevyeExcel.DATA
ws['A1'].number_format
ws.merge_cells('A2:A3')
ws.merge_cells('B2:B3')
ws.merge_cells('C2:E2')
ws['A2'] = 'Продукция'
ws['B2'] = 'Цена'
ws['C2'] = 'Изменения относительно предыдущего периода'
ws['C3'] = 'День'
ws['D3'] = 'Неделя'
ws['E3'] = 'Месяц'
ws['A4'] = 'Железнорудное сырье'
ws['A5'] = SurevyeExcel.cellFE
ws['A6'] = 'Уголь'
ws['A7'] = SurevyeExcel.cellUg
ws['A8'] = 'Кокс'
ws['A9'] = SurevyeExcel.cellKo
ws['A10'] = 'Лом'
ws['A11'] = SurevyeExcel.cellLo
ws['A12'] = SurevyeExcel.cellRu
ws['A13'] = 'Губчатое железо'
ws['A14'] = SurevyeExcel.cellIn
ws['A15'] = 'Чугун'
ws['A16'] = SurevyeExcel.cellCh
ws['A17'] = SurevyeExcel.cellFob
ws['B5'] = str(SurevyeExcel.PRICE1) + " USD/т"
ws['B7'] = str(SurevyeExcel.PRICE2) + " USD/т"
ws['B9'] = str(SurevyeExcel.PRICE3) + " USD/т"
ws['B11'] = str(SurevyeExcel.PRICE4) + " USD/т"
ws['B12'] = str(SurevyeExcel.PRICE5) + " руб/т"
ws['B14'] = str(SurevyeExcel.PRICE6) + " USD/т"
ws['B16'] = str(SurevyeExcel.PRICE7) + " USD/т"
ws['B17'] = str(SurevyeExcel.PRICE8) + " USD/т"
ws['C5'] = f"{SurevyeExcel.CH_DAY1:+}" + '  (' + f"{SurevyeExcel.CH_DAY_PR1:+}" + " %)"
ws['C7'] = f"{SurevyeExcel.CH_DAY2:+}" + '  (' + f"{SurevyeExcel.CH_DAY_PR2:+}" + " %)"
ws['C9'] = f"{SurevyeExcel.CH_DAY3:+}" + '  (' + f"{SurevyeExcel.CH_DAY_PR3:+}" + " %)"
ws['C11'] = f"{SurevyeExcel.CH_DAY4:+}" + '  (' + f"{SurevyeExcel.CH_DAY_PR4:+}" + " %)"
ws['C12'] = f"{SurevyeExcel.CH_DAY5:+}" + '  (' + f"{SurevyeExcel.CH_DAY_PR5:+}" + " %)"
ws['C14'] = f"{SurevyeExcel.CH_DAY6:+}" + '  (' + f"{SurevyeExcel.CH_DAY_PR6:+}" + " %)"
ws['C16'] = f"{SurevyeExcel.CH_DAY7:+}" + '  (' + f"{SurevyeExcel.CH_DAY_PR7:+}" + " %)"
ws['C17'] = f"{SurevyeExcel.CH_DAY8:+}" + '  (' + f"{SurevyeExcel.CH_DAY_PR8:+}" + " %)"
ws['D5'] = f"{SurevyeExcel.CH_W1:+}" + '  (' + f"{SurevyeExcel.CH_W_PR1:+}" + " %)"
ws['D7'] = f"{SurevyeExcel.CH_W2:+}" + '  (' + f"{SurevyeExcel.CH_W_PR2:+}" + " %)"
ws['D9'] = f"{SurevyeExcel.CH_W3:+}" + '  (' + f"{SurevyeExcel.CH_W_PR3:+}" + " %)"
ws['D11'] = f"{SurevyeExcel.CH_W4:+}" + '  (' + f"{SurevyeExcel.CH_W_PR4:+}" + " %)"
ws['D12'] = f"{SurevyeExcel.CH_W5:+}" + '  (' + f"{SurevyeExcel.CH_W_PR5:+}" + " %)"
ws['D14'] = f"{SurevyeExcel.CH_W6:+}" + '  (' + f"{SurevyeExcel.CH_W_PR6:+}" + " %)"
ws['D16'] = f"{SurevyeExcel.CH_W7:+}" + '  (' + f"{SurevyeExcel.CH_W_PR7:+}" + " %)"
ws['D17'] = f"{SurevyeExcel.CH_W8:+}" + '  (' + f"{SurevyeExcel.CH_W_PR8:+}" + " %)"
ws['E5'] = f"{SurevyeExcel.CH_M1:+}" + '  (' + f"{SurevyeExcel.CH_M_PR1:+}" + " %)"
ws['E7'] = f"{SurevyeExcel.CH_M2:+}" + '  (' + f"{SurevyeExcel.CH_M_PR2:+}" + " %)"
ws['E9'] = f"{SurevyeExcel.CH_M3:+}" + '  (' + f"{SurevyeExcel.CH_M_PR3:+}" + " %)"
ws['E11'] = f"{SurevyeExcel.CH_M4:+}" + '  (' + f"{SurevyeExcel.CH_M_PR4:+}" + " %)"
ws['E12'] = f"{SurevyeExcel.CH_M5:+}" + '  (' + f"{SurevyeExcel.CH_M_PR5:+}" + " %)"
ws['E14'] = f"{SurevyeExcel.CH_M6:+}" + '  (' + f"{SurevyeExcel.CH_M_PR6:+}" + " %)"
ws['E16'] = f"{SurevyeExcel.CH_M7:+}" + '  (' + f"{SurevyeExcel.CH_M_PR7:+}" + " %)"
ws['E17'] = f"{SurevyeExcel.CH_M8:+}" + '  (' + f"{SurevyeExcel.CH_M_PR8:+}" + " %)"
def set_border(ws, cell_range):
    thin = Side(border_style="thin", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
set_border(ws, 'A1:E17') 
wb.save('Sur.xlsx')
wb = openpyxl.load_workbook('Sur.xlsx')
sheet = wb.active
custom_page_width = 1270 
custom_page_height = 380
custom_page_size = (custom_page_width, custom_page_height)
c = canvas.Canvas('Sur.pdf', pagesize=custom_page_size)
pdfmetrics.registerFont(TTFont('Arial', 'arial.ttf'))
num_rows = sheet.max_row
num_cols = sheet.max_column
cell_width = 250
cell_height = 20
x_start = 10
y_start = 360
for i in range(1, num_rows + 1):
    y = y_start - (i * cell_height)
    for j in range(1, num_cols + 1):
        x = x_start + ((j - 1) * cell_width)
        value = sheet.cell(row=i, column=j).value
        if isinstance(value, (float, int)):
            value = str(value)
        elif isinstance(value, datetime):
            value = value.strftime('%Y-%m-%d %H:%M:%S')
        elif value is None:
            value = ''
        c.setFont('Arial', 10)
        c.drawCentredString(x + (cell_width / 2), y + (cell_height / 2), value)
        c.rect(x, y, cell_width, cell_height)
    for i in range(num_rows + 1):
        y = y_start - (i * cell_height)
        c.line(x_start, y, x_start + (num_cols * cell_width), y)
    for j in range(num_cols + 1):
        x = x_start + (j * cell_width)
        c.line(x, y_start, x, y_start - (num_rows * cell_height))
c.save()
images = convert_from_path('Sur.pdf', 700,poppler_path = os.getenv('POPPLERSUPER'))
for image in images:
    image.save('Sur.png')