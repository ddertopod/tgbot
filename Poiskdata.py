import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import datetime
from datetime import datetime

with open("datetime.txt", "r") as file:
    saved_datetime = file.read()
saved_datetime = datetime.strptime(saved_datetime, '%d.%m.%Y')
print(saved_datetime)
print("выше")
def find_datetime_row(file_path, search_value):
    wb = load_workbook(file_path)
    sheet = wb.active
    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start = 1):
        for cell_value in row:
            if isinstance(cell_value, datetime) and cell_value == search_value:
                return row_number
    return -1
file_path = "Svodka.xlsx"
search_value = saved_datetime
row_number = find_datetime_row(file_path, search_value)
if row_number == -1:
    print("Искомое значение не найдено")
else:
    print(f"Номер строки с искомым значением: {row_number}")
NomerStroki = str(row_number)
print(NomerStroki)